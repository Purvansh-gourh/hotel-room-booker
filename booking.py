# features
# 1. search for a room
#     a. book if available
# 2. search for booked room details
# check-out

import openpyxl
import datetime
from datetime import date

wb = openpyxl.load_workbook('rooms.xlsx')
sheet = wb['Sheet1']
rec = openpyxl.load_workbook('record.xlsx')
record = rec['Sheet1']

while 1:
    br = 0
    # searching a room
    while 1:
        print('Enter choice')
        print('1. search for a room')
        print('2. search booked room details')
        print('3. check out')
        print('-1 to exit')
        choice = int(input())

        if choice == -1:
            br = 1
            break

        # input suit type
        if choice == 1:
            print('Enter suit choice')
            print('1. AC')
            print('2. Non Ac')
            print('3. Deluxe')

            roomtype = int(input())

            roomtypestr = ''
            price = 0
            if roomtype == 1:
                roomtypestr = 'AC'
                price = 2000
            elif roomtype == 2:
                roomtypestr = 'NonAc'
                price = 1200
            elif roomtype == 3:
                roomtypestr = 'Deluxe'
                price = 4000
            else:
                print('enter valid type')
                break

            # search for availability from excel file
            rowval = 2
            while sheet.cell(row=rowval, column=3).value != roomtypestr:
                rowval += 1
            x = rowval
            for i in range(x, x + 4):
                if sheet.cell(row=rowval, column=2).value == 'available':
                    break
                rowval += 1
            if rowval == x + 4:
                print('No rooms available in this suit')
                break
            print('Room available')
            x = input('Continue to booking : y/n :-')
            if x != 'y':
                br = 0
                break

            # input details
            name = input('Enter name : ')
            while 1:
                mob = int(input('Enter 10 digit mobile number : '))
                if mob < 1000000000 or mob > 9999999999:
                    print('invalid mobile number')
                else:
                    break

            address = input('Enter address : ')

            while 1:

                # today = date.today()
                checkin_str = input('enter check in date in dd/mm/yy : ')
                checkout_str = input('enter check out date in dd/mm/yy : ')

                checkin = datetime.datetime.strptime(checkin_str, '%d/%m/%Y')
                checkout = datetime.datetime.strptime(checkout_str, '%d/%m/%Y')

                # if checkin < today or checkout <today:
                #     print('check-in/check-out in past !! Not allowed here')

                if checkin != checkout:
                    stay_time_str = str(checkout - checkin)
                    x = stay_time_str.split(' ')
                    stay_time = int(x[0])
                else:
                    stay_time = 1

                if stay_time < 0:
                    print('Invalid check in date .re-enter dates')
                elif stay_time > 60:
                    print('booking for more than two months not allowed . re-enter dates')
                else:
                    break

            total_amount = stay_time * price

            print('total amount : ', total_amount)
            advance = int(input('Enter advance : '))
            net_remain = total_amount - advance

            # output bill and details
            print('--------:::::::::::: Welcome ::::::::::::--------')
            print('Details :')
            print('name :', name)
            print('room no : ', rowval - 1)
            print('room type : ', roomtypestr)
            print('check in : ', checkin)
            print('check out : ', checkout)
            print('address : ', address)
            print('mob : ', mob)
            print('total amount : ', total_amount)
            print('advance paid : ', advance)
            print('net remaining : ', net_remain)
            print('\nthank you')

            # enter  details to file
            sheet.cell(row=rowval, column=2).value = 'booked'
            sheet.cell(row=rowval, column=4).value = name
            sheet.cell(row=rowval, column=5).value = checkin
            sheet.cell(row=rowval, column=5).number_format = 'DD/MM/YYYY'
            sheet.cell(row=rowval, column=6).value = checkout
            sheet.cell(row=rowval, column=6).number_format = 'DD/MM/YYYY'
            sheet.cell(row=rowval, column=7).value = mob
            sheet.cell(row=rowval, column=7).number_format = 'number'
            sheet.cell(row=rowval, column=8).value = address
            sheet.cell(row=rowval, column=9).value = total_amount
            # sheet.cell(row=rowval, column=9).number_format = '#,##0.00'
            sheet.cell(row=rowval, column=10).value = advance
            # sheet.cell(row=rowval, column=10).number_format = 'number'
            sheet.cell(row=rowval, column=11).value = net_remain
            # sheet.cell(row=rowval, column=11).number_format = 'number'

            print('room booked succcesfully')
            wb.save('rooms.xlsx')

            # entering into record file
            last_row = record.max_row
            if 1 == last_row:
                entry_no = 1
            else:
                entry_no = record.cell(row=last_row, column=1).value+1
            last_row += 1
            record.cell(row=last_row, column=1).value = entry_no
            record.cell(row=last_row, column=2).value = rowval - 1
            record.cell(row=last_row, column=3).value = roomtypestr
            record.cell(row=last_row, column=4).value = name
            record.cell(row=last_row, column=5).value = checkin
            record.cell(row=last_row, column=5).number_format = 'DD/MM/YYYY'
            record.cell(row=last_row, column=6).value = checkout
            record.cell(row=last_row, column=6).number_format = 'DD/MM/YYYY'
            record.cell(row=last_row, column=7).value = mob
            record.cell(row=last_row, column=7).number_format = 'number'
            record.cell(row=last_row, column=8).value = address
            record.cell(row=last_row, column=9).value = total_amount
            rec.save('record.xlsx')

            br = abs(int(input('-1 to main menu')))

        # search for room booked by name
        if choice == 2:
            name = input('Enter name to be searched')
            rowval = 2
            found = 0
            for i in range(rowval, rowval + 12):
                name_list = str(sheet.cell(row=i, column=4).value)
                name_list = name_list.split()
                if name in name_list:
                    print('room found')
                    print('room no :', i - 1)
                    print('check in : ', sheet.cell(row=i, column=5).value)
                    print('check out : ', sheet.cell(row=i, column=6).value)
                    found = 1
                    break
            if found == 0:
                print('No room booked with this name')

            br = abs(int(input('-1 to main menu')))

        if choice == 3:
            room_no = int(input('Enter your room no : '))
            rowval = room_no + 1

            if sheet.cell(row=rowval, column=2).value != 'booked':
                print('This room is not booked')
                br = 0
                break
            print('Hello ', sheet.cell(row=rowval, column=4).value)
            print('total amount : ', sheet.cell(row=rowval, column=9).value)
            print('Net amount to be paid : ', sheet.cell(row=rowval, column=11).value)
            print('advance paid : ', sheet.cell(row=rowval, column=10).value)
            print('Thanks for visiting.....')
            sheet.cell(row=rowval, column=2).value = 'available'
            for i in range(4, 12):
                sheet.cell(row=rowval, column=i).value = ' '

            br = abs(int(input('-1 to main menu')))
            wb.save('rooms.xlsx')

    if br == 1:
        break
wb.save('rooms.xlsx')
